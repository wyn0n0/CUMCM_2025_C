import sys
from typing import Optional, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:
    sys.stderr.write(
        "Missing dependency 'openpyxl'. Please install it with: python -m pip install --user openpyxl\n"
    )
    raise


def find_column_index_by_header(sheet: "Worksheet", header_name: str) -> Optional[int]:
    if sheet.max_row < 1:
        return None
    for column_index, cell in enumerate(sheet[1], start=1):
        value = cell.value
        if isinstance(value, str):
            normalized = value.strip()
        else:
            normalized = str(value).strip() if value is not None else ""
        if normalized == header_name:
            return column_index
    return None


def get_or_create_column_index(sheet: "Worksheet", header_name: str) -> int:
    existing_index = find_column_index_by_header(sheet, header_name)
    if existing_index is not None:
        return existing_index
    new_index = sheet.max_column + 1
    sheet.cell(row=1, column=new_index, value=header_name)
    return new_index


def is_empty_cell_value(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def update_sheet(file_path: str, sheet_name: str) -> Tuple[int, int, int]:
    workbook = load_workbook(filename=file_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in '{file_path}'. Available: {workbook.sheetnames}")

    sheet: Worksheet = workbook[sheet_name]

    source_header = "染色体的非整数倍"
    target_header = "胎儿异常"

    source_col_index = find_column_index_by_header(sheet, source_header)
    if source_col_index is None:
        raise ValueError(
            f"Column '{source_header}' not found in first row of sheet '{sheet_name}'."
        )

    target_col_index = get_or_create_column_index(sheet, target_header)

    zeros_written = 0
    ones_written = 0
    processed_rows = 0

    for row_index in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_index, column=source_col_index).value
        if is_empty_cell_value(cell_value):
            sheet.cell(row=row_index, column=target_col_index, value=0)
            zeros_written += 1
        else:
            sheet.cell(row=row_index, column=target_col_index, value=1)
            ones_written += 1
        processed_rows += 1

    workbook.save(file_path)
    return processed_rows, zeros_written, ones_written


def main(argv: list[str]) -> int:
    # Defaults for this project
    file_path = "附件.xlsx"
    sheet_name = "女胎检测数据"

    # Minimal CLI: allow overriding via args: file_path [sheet_name]
    if len(argv) >= 2:
        file_path = argv[1]
    if len(argv) >= 3:
        sheet_name = argv[2]

    try:
        processed_rows, zeros, ones = update_sheet(file_path, sheet_name)
        sys.stdout.write(
            (
                f"Updated '{file_path}' -> sheet '{sheet_name}'.\n"
                f"Processed rows (excluding header): {processed_rows}.\n"
                f"Wrote '{"胎儿异常"}' values: 0 => {zeros}, 1 => {ones}.\n"
            )
        )
        return 0
    except Exception as exc:  # noqa: BLE001 - surface readable error
        sys.stderr.write(f"Error: {exc}\n")
        return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))


